# TAC Semantic Mapping System
# Copyright (c) 2026 [Victor Yan Huang]
# Licensed under the MIT License

from flask import Flask, render_template, request, redirect, url_for, send_file, jsonify
import pandas as pd
import os
import json
from fuzzywuzzy import fuzz
from werkzeug.utils import secure_filename
from datetime import datetime
import time
import jwt
import threading
from watchdog.observers import Observer
from watchdog.events import FileSystemEventHandler

# Ensure log directories exist
log_folders = ['Update Knowledge Base log', 'Upload and Process log']
for folder in log_folders:
    if not os.path.exists(folder):
        os.makedirs(folder)
        print(f"Created directory: {folder}")

# Log file generation function
def generate_log_file(log_data, log_type):
    """Generate log file and save to specified directory"""
    # Determine log directory
    if log_type == 'update':
        log_folder = 'Update Knowledge Base log'
        log_prefix = 'Update Knowledge Base'
    else:
        log_folder = 'Upload and Process log'
        log_prefix = 'Upload and Process'
    
    # Generate timestamp
    timestamp = datetime.now().strftime('%Y%m%d%H%M%S%f')[:-3]
    log_filename = os.path.join(log_folder, f'{log_prefix}_{timestamp}.log')
    
    # Write log
    with open(log_filename, 'w', encoding='utf-8') as f:
        f.write(f"Log generated at: {datetime.now().strftime('%Y-%m-%d %H:%M:%S.%f')[:-3]}\n")
        f.write("=" * 80 + "\n")
        
        if log_type == 'update':
            # Update Knowledge Base log
            f.write("Update Knowledge Base Operation\n")
            f.write("=" * 80 + "\n")
            
            if 'read_time' in log_data:
                f.write("Processing Time:\n")
                f.write(f"  Reading time: {log_data['read_time']:.4f} seconds\n")
            
            if 'vectorize_time' in log_data:
                f.write(f"  Vectorization time: {log_data['vectorize_time']:.4f} seconds\n")
            
            if 'total_time' in log_data:
                f.write(f"  Total processing time: {log_data['total_time']:.4f} seconds\n")
            
            if 'record_count' in log_data:
                f.write("\nData Information:\n")
                f.write(f"  Total records: {log_data['record_count']}\n")
                f.write(f"  Processed items: {log_data['processed_count']}\n")
            
            if 'error_count' in log_data:
                f.write(f"  Error count: {log_data['error_count']}\n")
        
        else:
            # Upload and Process log
            f.write("Upload and Process Operation\n")
            f.write("=" * 80 + "\n")
            
            if 'read_time' in log_data:
                f.write("Processing Time:\n")
                f.write(f"  Reading time: {log_data['read_time']:.4f} seconds\n")
            
            if 'vectorize_time' in log_data:
                f.write(f"  Vectorization time: {log_data['vectorize_time']:.4f} seconds\n")
            
            if 'total_time' in log_data:
                f.write(f"  Total processing time: {log_data['total_time']:.4f} seconds\n")
            
            if 'record_count' in log_data:
                f.write("\nData Information:\n")
                f.write(f"  Total records: {log_data['record_count']}\n")
                f.write(f"  Processed items: {log_data['processed_count']}\n")
            
            if 'matching_info' in log_data:
                f.write("\nMatching Information:\n")
                f.write(f"  Exact matches: {log_data['matching_info'].get('exact_matches', 0)}\n")
                f.write(f"  Semantic matches: {log_data['matching_info'].get('semantic_matches', 0)}\n")
                f.write(f"  Fuzzy matches: {log_data['matching_info'].get('fuzzy_matches', 0)}\n")
                f.write(f"  No matches: {log_data['matching_info'].get('no_matches', 0)}\n")
            
            if 'vector_distances' in log_data and log_data['vector_distances']:
                f.write("\nVector Angular Similarities (Top 10):\n")
                f.write("  Rank  | Similarity | Master Sub Material | Import Sub Material\n")
                f.write("  " + "-" * 90 + "\n")
                for i, (similarity, master_sub, import_sub) in enumerate(log_data['vector_distances'][:10]):
                    f.write(f"  {i+1:4d} | {similarity:.4f} | {master_sub[:30]:30} | {import_sub[:30]:30}\n")
        
        # Write other information
        if 'other_info' in log_data:
            f.write("\nOther Information:\n")
            for key, value in log_data['other_info'].items():
                f.write(f"  {key}: {value}\n")
        
        f.write("=" * 80 + "\n")
    
    print(f"Log file generated: {log_filename}")
    return log_filename

# Try to import Annoy library for approximate nearest neighbor search
try:
    from annoy import AnnoyIndex
    use_ann = True
    print("Successfully imported Annoy library")
except ImportError:
    use_ann = False
    print("Annoy library not available, will use linear search")

# Try to load sentence-transformers and numpy, use fuzzy matching as fallback if unavailable
use_semantic_matching = False
model = None
np = None

print("Attempting to initialize semantic matching...")

try:
    import os
    import time
    
    # Try to import numpy first
    print("Attempting to import numpy...")
    import numpy as np
    print(f"Successfully imported numpy version: {np.__version__}")
    
    # Ensure models directory exists
    if not os.path.exists('models'):
        os.makedirs('models')
        print("Created models directory")
    
    # Try to import sentence-transformers
    print("Attempting to import sentence-transformers...")
    from sentence_transformers import SentenceTransformer
    print("Successfully imported sentence-transformers")
    
    # Model local storage path
    MODEL_DIR = 'models/all-MiniLM-L6-v2'
    
    # Try to load model
    try:
        print("Attempting to load model...")
        
        # First check if model exists locally
        if os.path.exists(MODEL_DIR) and os.path.isdir(MODEL_DIR):
            print(f"Loading model from local directory: {MODEL_DIR}")
            model = SentenceTransformer(MODEL_DIR)
            print("Model loaded successfully from local directory")
            use_semantic_matching = True
        else:
            # Not found locally, download from remote
            print("Model not found locally, downloading from Hugging Face...")
            model = SentenceTransformer('all-MiniLM-L6-v2')
            print("Model downloaded successfully")
            # Save model to local directory
            print("Saving model to local directory...")
            model.save(MODEL_DIR)
            print("Model saved successfully")
            use_semantic_matching = True
    except Exception as e:
        print(f"Error loading model: {str(e)}")
        import traceback
        traceback.print_exc()
        use_semantic_matching = False
    
except ImportError as ie:
    print(f"Import error: {str(ie)}")
    print("Warning: sentence-transformers or numpy not available. Using fuzzy matching only.")
    use_semantic_matching = False
except Exception as e:
    print(f"Error loading semantic matching components: {str(e)}")
    import traceback
    traceback.print_exc()
    use_semantic_matching = False

print(f"Semantic matching enabled: {use_semantic_matching}")
print(f"Model object: {model}")

# Global variables for storing preprocessed data
master_dict = {}
master_items = []  # Store master file items and values
master_embeddings = None  # Store master file embeddings
master_item_to_idx = {}  # Store mapping from master item key to index for fast lookup
ann_index = None  # Store ANN index

# Mock user database for authentication
# Note: In production, use a proper user database with hashed passwords
users = {
    'testuser1': 'maersktac901',  # Example user
    'testuser2': 'maersktac902',  # Example user
    'testuser3': 'maersktac903',  # Example user
    'testuser4': 'maersktac904'   # Example user
}

# Token expiration time (seconds)
TOKEN_EXPIRATION = 72000

# Preprocess master file embeddings and cache
def preprocess_master_file():
    """Preprocess master file, calculate and cache embeddings"""
    global master_dict, master_embeddings, master_items, master_item_to_idx
    
    if not os.path.exists(MASTER_FILE_PATH):
        print("Master file not found, skipping preprocessing")
        return
    
    print("Preprocessing master file...")
    
    # Record start time
    start_time = time.time()
    
    df_master = pd.read_excel(MASTER_FILE_PATH)
    
    # Record reading completion time
    read_time = time.time()
    read_duration = read_time - start_time
    
    # Create master file dictionary
    master_dict = {(row['MATERIAL_NAME'], row['SUB_MATERIAL_NAME'], row['UOM']): 
                   (row['TRANSACTION_TYPE'], row['IMLUI_Billable Date'], row['IMLUI_Billable Qty'], row['IMLUI_Document ID']) 
                   for _, row in df_master.iterrows() if pd.notna(row['UOM'])}
    
    # Preprocess embeddings (only when semantic matching is enabled)
    vectorize_duration = 0
    error_count = 0
    
    if use_semantic_matching and model:
        master_items = []
        master_embeddings = []
        master_item_to_idx = {}  # Reset the mapping
        
        # Record vectorization start time
        vectorize_start_time = time.time()
        
        for idx, ((material_name, sub_material_name, uom), values) in enumerate(master_dict.items()):
            master_items.append(((material_name, sub_material_name, uom), values))
            # Create mapping from master key to index for fast lookup
            master_item_to_idx[(material_name, sub_material_name, uom)] = idx
            try:
                # Generate embedding for SUB_MATERIAL_NAME
                embedding = model.encode(str(sub_material_name))
                master_embeddings.append(embedding)
            except Exception as e:
                print(f"Error encoding {sub_material_name}: {str(e)}")
                error_count += 1
                if model:
                    # Create zero vector as fallback
                    embedding_dim = model.get_sentence_embedding_dimension()
                    master_embeddings.append(np.zeros(embedding_dim))
                else:
                    master_embeddings.append(np.zeros(384))  # Default dimension
        
        # Record vectorization completion time
        vectorize_end_time = time.time()
        vectorize_duration = vectorize_end_time - vectorize_start_time
        
        # Convert embeddings to numpy array for faster calculation
        if master_embeddings:
            master_embeddings = np.array(master_embeddings)
            print(f"Preprocessed {len(master_embeddings)} embeddings for master file")
            print(f"Master file vectorization time: {vectorize_duration:.4f} seconds")
    
    # Record preprocessing completion time
    end_time = time.time()
    total_duration = end_time - start_time
    
    print(f"Master file reading time: {read_duration:.4f} seconds")
    print(f"Total master file preprocessing time: {total_duration:.4f} seconds")
    print("Master file preprocessing completed")
    
    # Generate log
    log_data = {
        'read_time': read_duration,
        'vectorize_time': vectorize_duration,
        'total_time': total_duration,
        'record_count': len(df_master),
        'processed_count': len(master_dict),
        'error_count': error_count,
        'other_info': {
            'use_semantic_matching': use_semantic_matching,
            'model_available': model is not None,
            'master_file_path': MASTER_FILE_PATH
        }
    }
    generate_log_file(log_data, 'update')

# Initialize ANN index
def init_ann_index():
    """Initialize approximate nearest neighbor index"""
    global ann_index
    
    if not use_semantic_matching or not use_ann or master_embeddings is None or len(master_embeddings) == 0:
        print("Skipping ANN index initialization")
        return
    
    try:
        dimension = master_embeddings.shape[1]
        ann_index = AnnoyIndex(dimension, 'angular')  # Use angular distance (suitable for cosine similarity)
        
        for i, embedding in enumerate(master_embeddings):
            ann_index.add_item(i, embedding)
        
        # Build index, more trees mean higher accuracy but longer build time
        ann_index.build(10)
        print(f"ANN index built successfully with {len(master_embeddings)} items")
    except Exception as e:
        print(f"Error initializing ANN index: {str(e)}")
        ann_index = None


app = Flask(__name__)
app.config['UPLOAD_FOLDER'] = 'uploads'
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024  # 16MB max size
app.secret_key = 'your_secret_key_here'  # Change this to a secure key

# JWT configuration for token validation and generation
# Note: In production, use environment variable for SECRET_KEY
app.config['SECRET_KEY'] = 'your-secret-key-here'  # Secret key for JWT token signing


# def generate_token(username):
#     """
#     Generate JWT token for authenticated user
#     This function can be replaced with other token generation methods in the future
#     """
#     import datetime
#     payload = {
#         'username': username,
#         'exp': datetime.datetime.utcnow() + datetime.timedelta(seconds=TOKEN_EXPIRATION),
#         'iat': datetime.datetime.utcnow()
#     }
#     return jwt.encode(payload, app.config['SECRET_KEY'], algorithm='HS256')


# def validate_token(token):
#     """
#     Validate JWT token
#     This function can be replaced with other token validation methods in the future
#     """
#     try:
#         payload = jwt.decode(token, app.config['SECRET_KEY'], algorithms=['HS256'])
#         return payload['username']
#     except jwt.ExpiredSignatureError:
#         return None
#     except jwt.InvalidTokenError:
#         return None


# def require_auth(f):
#     """
#     Decorator to require authentication for API endpoints
#     This decorator can be modified to use other authentication methods in the future
#     """
#     def decorated(*args, **kwargs):
#         auth_header = request.headers.get('Authorization')
#         if not auth_header:
#             return jsonify({'error': 'Authorization header is required'}), 401
#         
#         try:
#             token = auth_header.split(' ')[1]  # Extract Bearer token
#         except IndexError:
#             return jsonify({'error': 'Invalid authorization header format'}), 401
#         
#         if not validate_token(token):
#             return jsonify({'error': 'Invalid or expired token'}), 401
#         
#         return f(*args, **kwargs)
#     decorated.__name__ = f.__name__
#     return decorated

# Path to the master Excel file (Dataset A)
# Separate directories for knowledge base, prediction files, and results
KB_FOLDER = 'knowledge_base'
RESULTS_FOLDER = 'results'
MASTER_FILE_PATH = os.path.join(KB_FOLDER, 'A.xlsx')

# Allowed file extensions
ALLOWED_EXTENSIONS = {'xlsx'}

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

def process_excel(file_path):
    """Process the uploaded Excel file (Dataset B) against the master file (Dataset A)"""
    try:
        # Record processing start time
        process_start_time = time.time()
        
        # Load master file (Dataset A)
        if not os.path.exists(MASTER_FILE_PATH):
            return None, "Master file (A.xlsx) not found. Please upload it first."
        
        # Ensure master file data is preprocessed
        if not master_dict:
            preprocess_master_file()
        
        # Record reading start time
        read_start_time = time.time()
        df_user = pd.read_excel(file_path)
        
        # Record reading completion time
        read_end_time = time.time()
        read_duration = read_end_time - read_start_time
        
        # Check if required columns exist
        required_columns = ['MATERIAL_NAME', 'SUB_MATERIAL_NAME', 'UOM', 'TRANSACTION_TYPE', 'IMLUI_Billable Date', 'IMLUI_Billable Qty', 'IMLUI_Document ID']
        if not all(col in df_user.columns for col in required_columns):
            return None, "Uploaded file is missing required columns."
        
        # Convert columns to string type to avoid dtype compatibility issues
        string_columns = ['TRANSACTION_TYPE', 'IMLUI_Billable Date', 'IMLUI_Billable Qty', 'IMLUI_Document ID']
        for col in string_columns:
            df_user[col] = df_user[col].astype(str).replace('nan', '')
        
        # Use globally preprocessed master_dict
        
        # Dictionary to track the matching method used for each row
        matching_methods = {}
        # Dictionary to track the best match scores for each row
        best_match_scores = {}
        # Dictionary to track semantic scores for each row
        semantic_scores = {}
        # Dictionary to track fuzzy scores for each row
        fuzzy_scores = {}
        # Dictionary to track matched master_sub_material for each row
        matched_master_sub_materials = {}
        # Record vector distance information
        vector_distances = []
        # Record matching information
        matching_info = {
            'exact_matches': 0,
            'semantic_matches': 0,
            'fuzzy_matches': 0,
            'no_matches': 0
        }
        # Record vectorization start time
        vectorize_start_time = time.time()
        
        # Process each row in user file
        total_rows = len(df_user)
        print(f"Starting to process {total_rows} rows...")
        
        for index, row in df_user.iterrows():
            # Progress feedback every 10 rows or at the end
            if (index + 1) % 10 == 0 or (index + 1) == total_rows:
                progress = ((index + 1) / total_rows) * 100
                print(f"Processing row {index + 1}/{total_rows} ({progress:.1f}%)")
            
            # Check if MATERIAL_NAME, SUB_MATERIAL_NAME and UOM have values
            if pd.notna(row['MATERIAL_NAME']) and pd.notna(row['SUB_MATERIAL_NAME']) and pd.notna(row['UOM']):
                material_name = row['MATERIAL_NAME']
                sub_material_name = row['SUB_MATERIAL_NAME']
                uom = row['UOM']
                
                # Check for exact match first
                if (material_name, sub_material_name, uom) in master_dict:
                    transaction_type, billable_date, billable_qty, document_id = master_dict[(material_name, sub_material_name, uom)]
                    df_user.at[index, 'TRANSACTION_TYPE'] = transaction_type
                    df_user.at[index, 'IMLUI_Billable Date'] = billable_date
                    df_user.at[index, 'IMLUI_Billable Qty'] = billable_qty
                    df_user.at[index, 'IMLUI_Document ID'] = document_id
                    matching_info['exact_matches'] += 1
                else:
                    # Perform semantic matching using fuzzy string matching
                    best_match = None
                    best_score = 0
                    best_matching_method = 'FuzzyMapping'
                    best_semantic_score = 0
                    best_fuzzy_score = 0
                    best_master_sub_material = ''
                    
                    # Generate user input embedding (only once)
                    user_embedding = None
                    if use_semantic_matching and model:
                        try:
                            user_embedding = model.encode(str(sub_material_name))
                        except Exception as e:
                            print(f"Error encoding user input: {str(e)}")
                    
                    # Use ANN index for fast nearest neighbor search
                    if use_semantic_matching and ann_index and user_embedding is not None:
                        try:
                            # Find nearest N matches
                            n_neighbors = min(10, len(master_items))  # Only find top 10 most similar
                            nearest_ids = ann_index.get_nns_by_vector(user_embedding, n_neighbors)
                            
                            # Detailed scoring for found neighbors
                            for neighbor_id in nearest_ids:
                                (master_material, master_sub_material, master_uom), values = master_items[neighbor_id]
                                
                                # Calculate combined similarity score
                                # For MATERIAL_NAME and UOM, use exact match or character similarity
                                material_score = fuzz.token_sort_ratio(str(material_name), str(master_material))
                                uom_score = fuzz.token_sort_ratio(str(uom), str(master_uom))
                                
                                # For SUB_MATERIAL_NAME, use both semantic similarity and fuzzy matching
                                current_fuzzy_score = fuzz.token_sort_ratio(str(sub_material_name), str(master_sub_material))
                                current_semantic_score = 0
                                
                                # Calculate semantic similarity using preprocessed embeddings
                                try:
                                    embedding2 = master_embeddings[neighbor_id]
                                    # Calculate cosine similarity
                                    dot_product = np.dot(user_embedding, embedding2)
                                    norm1 = np.linalg.norm(user_embedding)
                                    norm2 = np.linalg.norm(embedding2)
                                    current_semantic_score = (dot_product / (norm1 * norm2)) * 100 if (norm1 * norm2) > 0 else 0
                                    # Use weighted average of semantic and fuzzy scores
                                    sub_material_score = (current_semantic_score * 0.9 + current_fuzzy_score * 0.1)
                                    matching_method = 'AISemanticMapping'  # Using semantic mapping
                                except Exception as e:
                                    print(f"Error in semantic matching: {str(e)}")
                                    # Fallback to fuzzy matching only if semantic similarity fails
                                    sub_material_score = current_fuzzy_score
                                    matching_method = 'FuzzyMapping'  # Using only fuzzy mapping
                                
                                # Give more weight to SUB_MATERIAL_NAME as it's the most important for semantic matching
                                combined_score = (material_score * 0.2 + sub_material_score * 0.6 + uom_score * 0.2)
                                
                                if combined_score > best_score:
                                    best_score = combined_score
                                    best_match = values
                                    best_matching_method = matching_method
                                    best_semantic_score = current_semantic_score
                                    best_fuzzy_score = current_fuzzy_score
                                    best_master_sub_material = master_sub_material
                        except Exception as e:
                            print(f"Error using ANN index: {str(e)}")
                            # Fallback to linear search
                            pass
                    
                    # If ANN index is unavailable or error occurs, fallback to linear search
                    if best_match is None:
                        # Search all master items
                        limited_master_items = list(master_dict.items())
                        
                        # OPTIMIZATION: Generate user embedding only once before the loop
                        user_embedding_linear = None
                        if use_semantic_matching and model and user_embedding is not None:
                            user_embedding_linear = user_embedding  # Reuse the already computed embedding
                        
                        for (master_material, master_sub_material, master_uom), values in limited_master_items:
                            # Calculate combined similarity score
                            # For MATERIAL_NAME and UOM, use exact match or character similarity
                            material_score = fuzz.token_sort_ratio(str(material_name), str(master_material))
                            uom_score = fuzz.token_sort_ratio(str(uom), str(master_uom))
                            
                            # For SUB_MATERIAL_NAME, use both semantic similarity and fuzzy matching
                            current_fuzzy_score = fuzz.token_sort_ratio(str(sub_material_name), str(master_sub_material))
                            current_semantic_score = 0
                            
                            if use_semantic_matching and model and user_embedding_linear is not None:
                                try:
                                    # OPTIMIZATION: Use pre-computed master embedding from master_embeddings array
                                    # Fast lookup using master_item_to_idx mapping (O(1) instead of O(n))
                                    master_key = (master_material, master_sub_material, master_uom)
                                    master_idx = master_item_to_idx.get(master_key)
                                    
                                    if master_idx is not None and master_idx < len(master_embeddings):
                                        embedding2 = master_embeddings[master_idx]
                                        # Calculate cosine similarity using pre-computed embeddings
                                        dot_product = np.dot(user_embedding_linear, embedding2)
                                        norm1 = np.linalg.norm(user_embedding_linear)
                                        norm2 = np.linalg.norm(embedding2)
                                        current_semantic_score = (dot_product / (norm1 * norm2)) * 100 if (norm1 * norm2) > 0 else 0
                                        # Use weighted average of semantic and fuzzy scores
                                        sub_material_score = (current_semantic_score * 0.9 + current_fuzzy_score * 0.1)
                                        matching_method = 'AISemanticMapping'  # Using semantic mapping
                                    else:
                                        # Fallback to fuzzy matching if embedding not found
                                        sub_material_score = current_fuzzy_score
                                        matching_method = 'FuzzyMapping'
                                except Exception as e:
                                    print(f"Error in semantic matching: {str(e)}")
                                    # Fallback to fuzzy matching only if semantic similarity fails
                                    sub_material_score = current_fuzzy_score
                                    matching_method = 'FuzzyMapping'  # Using only fuzzy matching
                            else:
                                # Use fuzzy matching only
                                sub_material_score = current_fuzzy_score
                                matching_method = 'FuzzyMapping'  # Using only fuzzy matching
                            
                            # Give more weight to SUB_MATERIAL_NAME as it's the most important for semantic matching
                            combined_score = (material_score * 0.2 + sub_material_score * 0.6 + uom_score * 0.2)
                            
                            if combined_score > best_score:
                                best_score = combined_score
                                best_match = values
                                best_matching_method = matching_method
                                best_semantic_score = current_semantic_score
                                best_fuzzy_score = current_fuzzy_score
                                best_master_sub_material = master_sub_material
                    
                    # Use the best match regardless of similarity score
                    if best_match:
                        transaction_type, billable_date, billable_qty, document_id = best_match
                        df_user.at[index, 'TRANSACTION_TYPE'] = transaction_type
                        df_user.at[index, 'IMLUI_Billable Date'] = billable_date
                        df_user.at[index, 'IMLUI_Billable Qty'] = billable_qty
                        df_user.at[index, 'IMLUI_Document ID'] = document_id
                        
                        # Store the matching method used for this row
                        matching_methods[index] = best_matching_method
                        best_match_scores[index] = best_score
                        semantic_scores[index] = best_semantic_score
                        fuzzy_scores[index] = best_fuzzy_score
                        matched_master_sub_materials[index] = best_master_sub_material
                        
                        # Record vector similarity
                        if best_semantic_score > 0:
                            vector_distances.append((best_semantic_score, best_master_sub_material, sub_material_name))
                        
                        if best_matching_method == 'AISemanticMapping':
                            matching_info['semantic_matches'] += 1
                        else:
                            matching_info['fuzzy_matches'] += 1
                    else:
                        matching_info['no_matches'] += 1
        
        # Record vectorization completion time
        vectorize_end_time = time.time()
        vectorize_duration = vectorize_end_time - vectorize_start_time
        
        # Record total processing time
        process_end_time = time.time()
        total_duration = process_end_time - process_start_time
        
        # Generate log
        log_data = {
            'read_time': read_duration,
            'vectorize_time': vectorize_duration,
            'total_time': total_duration,
            'record_count': len(df_user),
            'processed_count': len(df_user),
            'matching_info': matching_info,
            'vector_distances': vector_distances,
            'other_info': {
                'use_semantic_matching': use_semantic_matching,
                'model_available': model is not None,
                'ann_index_available': ann_index is not None
            }
        }
        generate_log_file(log_data, 'process')
        
        # Add new columns at the end
        df_user['Match Type'] = ''
        df_user['Matched Master SUB_MATERIAL_NAME'] = ''
        df_user['Semantic Score'] = ''
        df_user['Fuzzy Score'] = ''
        df_user['Final Combined Score'] = ''
        
        # Check each row and mark as exact match or semantic inference
        for index, row in df_user.iterrows():
            if pd.notna(row['MATERIAL_NAME']) and pd.notna(row['SUB_MATERIAL_NAME']) and pd.notna(row['UOM']):
                material_name = row['MATERIAL_NAME']
                sub_material_name = row['SUB_MATERIAL_NAME']
                uom = row['UOM']
                
                # Determine match type
                if (material_name, sub_material_name, uom) in master_dict:
                    df_user.at[index, 'Match Type'] = "Exact Match"
                    df_user.at[index, 'Matched Master SUB_MATERIAL_NAME'] = sub_material_name
                    df_user.at[index, 'Semantic Score'] = "-"
                    df_user.at[index, 'Fuzzy Score'] = "-"
                    df_user.at[index, 'Final Combined Score'] = "-"
                elif index in matching_methods:
                    # Use the matching method stored during processing
                    df_user.at[index, 'Match Type'] = matching_methods[index]
                    df_user.at[index, 'Matched Master SUB_MATERIAL_NAME'] = matched_master_sub_materials.get(index, "-")
                    # Convert numpy float types to strings
                    semantic_score = semantic_scores.get(index, "-")
                    fuzzy_score = fuzzy_scores.get(index, "-")
                    final_score = best_match_scores.get(index, "-")
                    
                    # Handle numpy float types
                    if semantic_score != "-":
                        semantic_score = str(semantic_score)
                    if fuzzy_score != "-":
                        fuzzy_score = str(fuzzy_score)
                    if final_score != "-":
                        final_score = str(final_score)
                    
                    df_user.at[index, 'Semantic Score'] = semantic_score
                    df_user.at[index, 'Fuzzy Score'] = fuzzy_score
                    df_user.at[index, 'Final Combined Score'] = final_score
                else:
                    df_user.at[index, 'Match Type'] = "Unknown"
                    df_user.at[index, 'Matched Master SUB_MATERIAL_NAME'] = "-"
                    df_user.at[index, 'Semantic Score'] = "-"
                    df_user.at[index, 'Fuzzy Score'] = "-"
                    df_user.at[index, 'Final Combined Score'] = "-"
        
        # Generate timestamp for output file name
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        
        # Save the result to a new file with timestamp in results directory
        output_file_path = os.path.join(RESULTS_FOLDER, f'MappingResult_{timestamp}.xlsx')
        
        # First save as regular Excel file
        df_user.to_excel(output_file_path, index=False)
        
        # Then open and format cells
        from openpyxl import load_workbook
        from openpyxl.styles import PatternFill
        
        wb = load_workbook(output_file_path)
        ws = wb.active
        
        # Find column indices
        header_row = 1
        final_score_col = None
        semantic_score_col = None
        sub_material_col = None
        matched_sub_material_col = None
        
        for col in ws.iter_cols(min_row=header_row, max_row=header_row):
            for cell in col:
                if cell.value == 'Final Combined Score':
                    final_score_col = cell.column
                elif cell.value == 'Semantic Score':
                    semantic_score_col = cell.column
                elif cell.value == 'SUB_MATERIAL_NAME':
                    sub_material_col = cell.column
                elif cell.value == 'Matched Master SUB_MATERIAL_NAME':
                    matched_sub_material_col = cell.column
        
        # Define fill styles
        red_fill = PatternFill(start_color='FF9999', end_color='FF9999', fill_type='solid')
        yellow_fill = PatternFill(start_color='FFFFCC', end_color='FFFFCC', fill_type='solid')
        
        # Format Final Combined Score cells below 70
        if final_score_col:
            for row in range(2, ws.max_row + 1):
                cell = ws.cell(row=row, column=final_score_col)
                try:
                    score = float(cell.value)
                    if score < 70:
                        cell.fill = red_fill
                except (ValueError, TypeError):
                    pass
        
        # Format Semantic Score cells below 50
        if semantic_score_col:
            for row in range(2, ws.max_row + 1):
                cell = ws.cell(row=row, column=semantic_score_col)
                try:
                    score = float(cell.value)
                    if score < 50:
                        cell.fill = red_fill
                except (ValueError, TypeError):
                    pass
        
        # Format SUB_MATERIAL_NAME cells with values
        if sub_material_col:
            for row in range(2, ws.max_row + 1):
                cell = ws.cell(row=row, column=sub_material_col)
                if cell.value and cell.value != '-':
                    cell.fill = yellow_fill
        
        # Format Matched Master SUB_MATERIAL_NAME cells with values
        if matched_sub_material_col:
            for row in range(2, ws.max_row + 1):
                cell = ws.cell(row=row, column=matched_sub_material_col)
                if cell.value and cell.value != '-':
                    cell.fill = yellow_fill
        
        # Save the formatted workbook
        wb.save(output_file_path)
        
        return output_file_path, None
        
    except Exception as e:
        return None, str(e)

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/upload', methods=['POST'])
def upload_file():
    if 'file' not in request.files:
        return redirect(request.url)
    
    file = request.files['file']
    
    if file.filename == '':
        return redirect(request.url)
    
    if file and allowed_file(file.filename):
        filename = secure_filename(file.filename)
        file_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
        file.save(file_path)
        
        # Process the file
        output_file, error = process_excel(file_path)
        
        if error:
            return render_template('index.html', error=error)
        
        # Extract the filename from the output_file path
        filename = os.path.basename(output_file)
        return send_file(output_file, as_attachment=True, download_name=filename)
    
    return redirect(request.url)

@app.route('/admin')
def admin():
    return render_template('admin.html')

@app.route('/admin/upload_master', methods=['POST'])
def upload_master():
    if 'file' not in request.files:
        return redirect(url_for('admin'))
    
    file = request.files['file']
    
    if file.filename == '':
        return redirect(url_for('admin'))
    
    if file and allowed_file(file.filename):
        # Ensure knowledge_base directory exists
        if not os.path.exists(KB_FOLDER):
            os.makedirs(KB_FOLDER)
        
        # Save the file as A.xlsx (overwrite if exists)
        file_path = MASTER_FILE_PATH
        file.save(file_path)
        
        # Re-preprocess master file data
        preprocess_master_file()
        # Re-initialize ANN index
        init_ann_index()
        
        return render_template('admin.html', message='Knowledge Base Update successfully!')
    
    return redirect(url_for('admin'))

def convert_numpy_types(obj):
    """
    Recursively convert numpy types to Python types for JSON serialization
    Handles all numpy scalar types and nested structures
    """
    import numpy as np
    if isinstance(obj, np.generic):  # Catch-all for all numpy scalar types
        return obj.item()  # Convert any numpy scalar to Python equivalent
    elif isinstance(obj, np.ndarray):
        return obj.tolist()
    elif isinstance(obj, dict):
        return {key: convert_numpy_types(value) for key, value in obj.items()}
    elif isinstance(obj, list):
        return [convert_numpy_types(item) for item in obj]
    elif isinstance(obj, tuple):
        return tuple(convert_numpy_types(item) for item in obj)
    else:
        return obj

# REST API interface
@app.route('/api/map', methods=['POST'])
# @require_auth
def api_map():
    """
    Semantic Mapping API Interface
    Receive JSON format request and return mapping results
    """
    try:
        # Get JSON data
        data = request.json
        if not data:
            return jsonify({'error': 'No JSON data provided'}), 400
        
        # Validate request data structure
        if 'items' not in data:
            return jsonify({'error': 'Missing "items" field'}), 400
        
        items = data['items']
        if not isinstance(items, list):
            return jsonify({'error': '"items" must be a list'}), 400
        
        # Process each item
        results = []
        for item in items:
            # Validate item structure
            if not all(key in item for key in ['MATERIAL_NAME', 'SUB_MATERIAL_NAME', 'UOM']):
                results.append({
                    'input': item,
                    'error': 'Missing required fields'
                })
                continue
            
            material_name = item['MATERIAL_NAME']
            sub_material_name = item['SUB_MATERIAL_NAME']
            uom = item['UOM']
            
            # Initialize result
            result = {
                'input': item,
                'output': {
                    'TRANSACTION_TYPE': '',
                    'IMLUI_Billable Date': '',
                    'IMLUI_Billable Qty': '',
                    'IMLUI_Document ID': '',
                    'Match Type': 'Unknown',
                    'Matched Master SUB_MATERIAL_NAME': '',
                    'Semantic Score': '-',
                    'Fuzzy Score': '-',
                    'Final Combined Score': '-'
                }
            }
            
            # Check for exact match
            if (material_name, sub_material_name, uom) in master_dict:
                transaction_type, billable_date, billable_qty, document_id = master_dict[(material_name, sub_material_name, uom)]
                result['output']['TRANSACTION_TYPE'] = transaction_type
                result['output']['IMLUI_Billable Date'] = billable_date
                result['output']['IMLUI_Billable Qty'] = billable_qty
                result['output']['IMLUI_Document ID'] = document_id
                result['output']['Match Type'] = 'Exact Match'
                result['output']['Matched Master SUB_MATERIAL_NAME'] = sub_material_name
            else:
                # Perform semantic and fuzzy matching
                best_match = None
                best_score = 0
                best_matching_method = 'FuzzyMapping'
                best_semantic_score = 0
                best_fuzzy_score = 0
                best_master_sub_material = ''
                
                # Generate user input embedding
                user_embedding = None
                if use_semantic_matching and model:
                    try:
                        user_embedding = model.encode(str(sub_material_name))
                    except Exception as e:
                        print(f"Error encoding user input: {str(e)}")
                
                # Use ANN index for fast lookup
                if use_semantic_matching and ann_index and user_embedding is not None:
                    try:
                        n_neighbors = min(10, len(master_items))
                        nearest_ids = ann_index.get_nns_by_vector(user_embedding, n_neighbors)
                        
                        for neighbor_id in nearest_ids:
                            (master_material, master_sub_material, master_uom), values = master_items[neighbor_id]
                            
                            # Calculate similarity scores
                            material_score = fuzz.token_sort_ratio(str(material_name), str(master_material))
                            uom_score = fuzz.token_sort_ratio(str(uom), str(master_uom))
                            current_fuzzy_score = fuzz.token_sort_ratio(str(sub_material_name), str(master_sub_material))
                            
                            # Calculate semantic similarity
                            current_semantic_score = 0
                            try:
                                embedding2 = master_embeddings[neighbor_id]
                                dot_product = np.dot(user_embedding, embedding2)
                                norm1 = np.linalg.norm(user_embedding)
                                norm2 = np.linalg.norm(embedding2)
                                current_semantic_score = (dot_product / (norm1 * norm2)) * 100 if (norm1 * norm2) > 0 else 0
                            except Exception as e:
                                print(f"Error in semantic matching: {str(e)}")
                            
                            # Calculate weighted score
                            sub_material_score = (current_semantic_score * 0.9 + current_fuzzy_score * 0.1)
                            combined_score = (material_score * 0.2 + sub_material_score * 0.6 + uom_score * 0.2)
                            
                            if combined_score > best_score:
                                best_score = combined_score
                                best_match = values
                                best_matching_method = 'AISemanticMapping'
                                best_semantic_score = current_semantic_score
                                best_fuzzy_score = current_fuzzy_score
                                best_master_sub_material = master_sub_material
                    except Exception as e:
                        print(f"Error in ANN search: {str(e)}")
                
                # If ANN search fails, use linear search
                if best_match is None:
                    limited_master_items = list(master_dict.items())
                    for (master_material, master_sub_material, master_uom), values in limited_master_items:
                        material_score = fuzz.token_sort_ratio(str(material_name), str(master_material))
                        uom_score = fuzz.token_sort_ratio(str(uom), str(master_uom))
                        current_fuzzy_score = fuzz.token_sort_ratio(str(sub_material_name), str(master_sub_material))
                        
                        current_semantic_score = 0
                        if use_semantic_matching and model:
                            try:
                                embedding1 = model.encode(str(sub_material_name))
                                embedding2 = model.encode(str(master_sub_material))
                                dot_product = np.dot(embedding1, embedding2)
                                norm1 = np.linalg.norm(embedding1)
                                norm2 = np.linalg.norm(embedding2)
                                current_semantic_score = (dot_product / (norm1 * norm2)) * 100 if (norm1 * norm2) > 0 else 0
                            except Exception as e:
                                print(f"Error in semantic matching: {str(e)}")
                        
                        sub_material_score = (current_semantic_score * 0.9 + current_fuzzy_score * 0.1) if use_semantic_matching else current_fuzzy_score
                        combined_score = (material_score * 0.2 + sub_material_score * 0.6 + uom_score * 0.2)
                        
                        if combined_score > best_score:
                            best_score = combined_score
                            best_match = values
                            best_matching_method = 'AISemanticMapping' if use_semantic_matching else 'FuzzyMapping'
                            best_semantic_score = current_semantic_score
                            best_fuzzy_score = current_fuzzy_score
                            best_master_sub_material = master_sub_material
                
                # Use best match regardless of score
                if best_match:
                    transaction_type, billable_date, billable_qty, document_id = best_match
                    result['output']['TRANSACTION_TYPE'] = transaction_type
                    result['output']['IMLUI_Billable Date'] = billable_date
                    result['output']['IMLUI_Billable Qty'] = billable_qty
                    result['output']['IMLUI_Document ID'] = document_id
                    result['output']['Match Type'] = best_matching_method
                    result['output']['Matched Master SUB_MATERIAL_NAME'] = best_master_sub_material
                    result['output']['Semantic Score'] = best_semantic_score if best_semantic_score > 0 else '-'
                    result['output']['Fuzzy Score'] = best_fuzzy_score
                    result['output']['Final Combined Score'] = best_score
            
            results.append(result)
        
        # Convert numpy types to Python types for JSON serialization
        results = convert_numpy_types(results)
        
        # Return results
        return jsonify({
            'status': 'success',
            'results': results,
            'metadata': {
                'total_items': len(items),
                'processed_items': len(results)
            }
        })
        
    except Exception as e:
        print(f"Error in API: {str(e)}")
        return jsonify({'error': str(e)}), 500

# API health check interface
@app.route('/api/health', methods=['GET'])
# @require_auth
def api_health():
    """
    API health check interface
    """
    return jsonify({
        'status': 'healthy',
        'services': {
            'semantic_matching': use_semantic_matching and model is not None,
            'ann_index': use_ann and 'ann_index' in globals() and globals()['ann_index'] is not None,
            'master_data': 'master_dict' in globals() and len(globals().get('master_dict', {})) > 0
        },
        'version': '1.0.0'
    })

# # Authentication endpoint for token generation
# # Note: This endpoint provides basic token generation for testing purposes
# # In production, consider replacing this with a more secure authentication system
# @app.route('/api/auth/login', methods=['POST'])
# def login():
#     """
#     User login to get access token
#     This endpoint can be replaced with other authentication methods in the future
#     """
#     data = request.json
#     if not data or 'username' not in data or 'password' not in data:
#         return jsonify({'error': 'Missing username or password'}), 400
#     
#     username = data['username']
#     password = data['password']
#     
#     # Validate user (use hashed password validation in production)
#     if username not in users or users[username] != password:
#         return jsonify({'error': 'Invalid credentials'}), 401
#     
#     # Generate token
#     token = generate_token(username)
#     return jsonify({
#         'access_token': token,
#         'token_type': 'Bearer',
#         'expires_in': TOKEN_EXPIRATION
#     })

# API documentation interface
@app.route('/api/docs', methods=['GET'])
# @require_auth
def api_docs():
    """
    API documentation interface
    """
    docs = {
        'endpoints': {
            '/api/map': {
                'method': 'POST',
                'description': 'Perform semantic mapping for multiple items',
                'request': {
                    'items': [{
                        'MATERIAL_NAME': 'string',
                        'SUB_MATERIAL_NAME': 'string',
                        'UOM': 'string'
                    }]
                },
                'response': {
                    'status': 'success',
                    'results': [{
                        'input': 'original input item',
                        'output': {
                            'TRANSACTION_TYPE': 'string',
                            'IMLUI_Billable Date': 'string',
                            'IMLUI_Billable Qty': 'string',
                            'IMLUI_Document ID': 'string',
                            'Match Type': 'string',
                            'Matched Master SUB_MATERIAL_NAME': 'string',
                            'Semantic Score': 'number or "-"',
                            'Fuzzy Score': 'number or "-"',
                            'Final Combined Score': 'number or "-"'
                        }
                    }],
                    'metadata': {
                        'total_items': 'number',
                        'processed_items': 'number'
                    }
                }
            },
            '/api/health': {
                'method': 'GET',
                'description': 'Check API health status'
            },
            '/api/docs': {
                'method': 'GET',
                'description': 'Get API documentation'
            }
        }
    }
    return jsonify(docs)


# File Watcher for automatic file processing
class FileWatcherHandler(FileSystemEventHandler):
    """Handler for file system events for knowledge base directory"""
    
    def __init__(self):
        self.last_processed = {}
        self.processing_lock = threading.Lock()
    
    def on_created(self, event):
        if event.is_directory:
            return
        
        file_path = event.src_path
        file_name = os.path.basename(file_path)
        
        # Only process Excel files
        if not file_name.lower().endswith(('.xlsx', '.xls')):
            return
        
        # Process all Excel files in knowledge_base directory (name not validated)
        self.process_knowledge_base_file(file_path)
    
    def on_modified(self, event):
        if event.is_directory:
            return
        
        file_path = event.src_path
        file_name = os.path.basename(file_path)
        
        # Only process Excel files
        if not file_name.lower().endswith(('.xlsx', '.xls')):
            return
        
        # Process all Excel files in knowledge_base directory (name not validated)
        self.process_knowledge_base_file(file_path)
    
    def process_knowledge_base_file(self, file_path):
        """Process knowledge base file"""
        # Prevent duplicate processing
        current_time = time.time()
        if file_path in self.last_processed and current_time - self.last_processed[file_path] < 2:
            return
        
        # Lock processing to prevent concurrency
        if not self.processing_lock.acquire(blocking=False):
            return
        
        try:
            # Wait 1 second to ensure file is fully written
            time.sleep(1)
            
            print("=" * 60)
            print(f"Detected new Knowledge Base file: {os.path.basename(file_path)}")
            print("=" * 60)
            
            # Copy to A.xlsx (overwrite existing file)
            import shutil
            if file_path != MASTER_FILE_PATH:
                shutil.copy2(file_path, MASTER_FILE_PATH)
                print(f"Copied {os.path.basename(file_path)} to {MASTER_FILE_PATH}")
            else:
                print(f"File is already at the correct location: {MASTER_FILE_PATH}")
            
            # Re-preprocess master file data
            preprocess_master_file()
            # Re-initialize ANN index
            init_ann_index()
            
            print("Knowledge Base file processing completed successfully!")
            print("=" * 60)
            
        except Exception as e:
            print(f"Error processing knowledge base file: {str(e)}")
        finally:
            self.processing_lock.release()
            self.last_processed[file_path] = current_time
    
    def on_created(self, event):
        """Handle file creation events"""
        if event.is_directory:
            return
        
        file_path = event.src_path
        file_name = os.path.basename(file_path)
        
        # Only process Excel files
        if not file_name.lower().endswith(('.xlsx', '.xls')):
            return
        
        # Wait a moment to ensure file is fully written
        time.sleep(1)
        
        # Process all Excel files in knowledge_base directory
        print(f"\n{'='*60}")
        print(f"Detected new Knowledge Base file: {file_name}")
        print(f"{'='*60}")
        self.process_knowledge_base_file(file_path)
    
    def on_modified(self, event):
        """Handle file modification events"""
        if event.is_directory:
            return
        
        file_path = event.src_path
        file_name = os.path.basename(file_path)
        
        # Only process Excel files
        if not file_name.lower().endswith(('.xlsx', '.xls')):
            return
        
        # Avoid processing the same file multiple times in quick succession
        current_time = time.time()
        with self.processing_lock:
            if file_path in self.last_processed:
                if current_time - self.last_processed[file_path] < 2:  # 2 seconds cooldown
                    return
            self.last_processed[file_path] = current_time
        
        # Wait a moment to ensure file is fully written
        time.sleep(1)
        
        print(f"\n{'='*60}")
        print(f"Detected modified Knowledge Base file: {file_name}")
        print(f"{'='*60}")
        self.process_knowledge_base_file(file_path)
    
    def process_knowledge_base_file(self, file_path):
        """Process Knowledge Base file and generate embeddings"""
        try:
            print(f"Processing Knowledge Base file: {file_path}")
            
            # Copy to A.xlsx (overwrite existing file)
            import shutil
            if file_path != MASTER_FILE_PATH:
                shutil.copy2(file_path, MASTER_FILE_PATH)
                print(f"Copied {os.path.basename(file_path)} to {MASTER_FILE_PATH}")
            else:
                print(f"File is already at the correct location: {MASTER_FILE_PATH}")
            
            # Preprocess master file
            preprocess_master_file()
            
            # Initialize ANN index
            init_ann_index()
            
            print(f"Knowledge Base file processing completed successfully!")
            print(f"{'='*60}\n")
            
        except Exception as e:
            print(f"Error processing Knowledge Base file: {str(e)}")
            import traceback
            traceback.print_exc()
            print(f"{'='*60}\n")


class PredictionFileHandler(FileSystemEventHandler):
    """Handle file system events for prediction files directory"""
    
    def __init__(self):
        self.last_processed = {}
        self.processing_lock = threading.Lock()
    
    def on_created(self, event):
        if event.is_directory:
            return
        
        file_path = event.src_path
        file_name = os.path.basename(file_path)
        
        # Only process Excel files
        if not file_name.lower().endswith(('.xlsx', '.xls')):
            return
        
        # Exclude system-generated result files
        if file_name.startswith('MappingResult_'):
            return
        
        # Process prediction file
        self.process_prediction_file(file_path)
    
    def on_modified(self, event):
        if event.is_directory:
            return
        
        file_path = event.src_path
        file_name = os.path.basename(file_path)
        
        # Only process Excel files
        if not file_name.lower().endswith(('.xlsx', '.xls')):
            return
        
        # Exclude system-generated result files
        if file_name.startswith('MappingResult_'):
            return
        
        # Process prediction file
        self.process_prediction_file(file_path)
    
    def process_prediction_file(self, file_path):
        """Process prediction file and generate output"""
        # Prevent duplicate processing
        current_time = time.time()
        if file_path in self.last_processed and current_time - self.last_processed[file_path] < 2:
            return
        
        # Lock processing to prevent concurrency
        if not self.processing_lock.acquire(blocking=False):
            return
        
        try:
            # Wait 1 second to ensure file is fully written
            time.sleep(1)
            
            print("=" * 60)
            print(f"Detected new Prediction file: {os.path.basename(file_path)}")
            print("=" * 60)
            
            # Process prediction file
            output_file, error = process_excel(file_path)
            
            if error:
                print(f"Error processing prediction file: {error}")
                print("=" * 60)
                return
            
            if output_file:
                print(f"Prediction file processed successfully!")
                print(f"Output file: {output_file}")
                print("=" * 60)
            else:
                print("Prediction file processing completed but no output file generated.")
                print("=" * 60)
            
        except Exception as e:
            print(f"Error processing prediction file: {str(e)}")
            import traceback
            traceback.print_exc()
            print("=" * 60)
        finally:
            self.processing_lock.release()
            self.last_processed[file_path] = current_time


def start_file_watcher():
    """Start file watchers for both knowledge base and prediction files directories"""
    try:
        # Ensure knowledge base directory exists
        if not os.path.exists(KB_FOLDER):
            os.makedirs(KB_FOLDER)
            print(f"Created {KB_FOLDER} directory for file watching")
        
        # Ensure uploads directory exists
        if not os.path.exists(app.config['UPLOAD_FOLDER']):
            os.makedirs(app.config['UPLOAD_FOLDER'])
            print(f"Created {app.config['UPLOAD_FOLDER']} directory for file watching")
        
        # Ensure results directory exists
        if not os.path.exists(RESULTS_FOLDER):
            os.makedirs(RESULTS_FOLDER)
            print(f"Created {RESULTS_FOLDER} directory for storing results")
        
        # Create observers and handlers for both directories
        kb_event_handler = FileWatcherHandler()
        kb_observer = Observer()
        kb_observer.schedule(kb_event_handler, KB_FOLDER, recursive=False)
        
        prediction_event_handler = PredictionFileHandler()
        prediction_observer = Observer()
        prediction_observer.schedule(prediction_event_handler, app.config['UPLOAD_FOLDER'], recursive=False)
        
        # Start observers in separate threads
        kb_observer.start()
        prediction_observer.start()
        
        print("\n" + "="*60)
        print("FILE WATCHERS STARTED")
        print("="*60)
        print(f"Knowledge Base directory: {os.path.abspath(KB_FOLDER)}")
        print(f"Prediction files directory: {os.path.abspath(app.config['UPLOAD_FOLDER'])}")
        print(f"Results directory: {os.path.abspath(RESULTS_FOLDER)}")
        print("Automatically processing files in both directories...")
        print("="*60 + "\n")
        
        # Store observers for cleanup
        global kb_file_observer, prediction_file_observer
        kb_file_observer = kb_observer
        prediction_file_observer = prediction_observer
        
    except Exception as e:
        print(f"Error starting file watchers: {str(e)}")
        print("File watchers not available. Manual upload through web interface still works.")


if __name__ == '__main__':
    # Ensure uploads directory exists
    if not os.path.exists(app.config['UPLOAD_FOLDER']):
        os.makedirs(app.config['UPLOAD_FOLDER'])
    
    # Ensure knowledge_base directory exists
    if not os.path.exists(KB_FOLDER):
        os.makedirs(KB_FOLDER)
    
    # Ensure results directory exists
    if not os.path.exists(RESULTS_FOLDER):
        os.makedirs(RESULTS_FOLDER)
    
    # Initialize preprocessing (continue even if master file doesn't exist)
    try:
        preprocess_master_file()
        # Initialize ANN index
        init_ann_index()
    except Exception as e:
        print(f"Error during initialization: {str(e)}")
        import traceback
        traceback.print_exc()
    
    # Performance Diagnostics
    print("\n" + "="*60)
    print("PERFORMANCE DIAGNOSTICS")
    print("="*60)
    print(f"Annoy library available: {use_ann}")
    print(f"Semantic matching enabled: {use_semantic_matching}")
    print(f"Model loaded: {model is not None}")
    if master_embeddings is not None:
        print(f"Master embeddings shape: {master_embeddings.shape}")
        print(f"Master embeddings count: {len(master_embeddings)}")
    else:
        print(f"Master embeddings: None")
    print(f"Master items count: {len(master_items)}")
    print(f"Master item index mapping: {len(master_item_to_idx)} entries")
    print(f"ANN index initialized: {ann_index is not None}")
    print(f"Master dict count: {len(master_dict)}")
    print("="*60 + "\n")
    
    # Start file watcher for uploads directory
    start_file_watcher()
    
    app.run(host='0.0.0.0', port=5000, debug=True)


# File Watcher for automatic file processing
class FileWatcherHandler(FileSystemEventHandler):
    """Handler for file system events"""
    def __init__(self):
        self.last_processed = {}
        self.processing_lock = threading.Lock()
    
    def on_created(self, event):
        """Handle file creation events"""
        if event.is_directory:
            return
        
        file_path = event.src_path
        file_name = os.path.basename(file_path)
        
        # Only process Excel files
        if not file_name.lower().endswith(('.xlsx', '.xls')):
            return
        
        # Wait a moment to ensure file is fully written
        time.sleep(1)
        
        # Check if file is A.xlsx (Knowledge Base file)
        if file_name == 'A.xlsx':
            print(f"\n{'='*60}")
            print(f"Detected new Knowledge Base file: {file_name}")
            print(f"{'='*60}")
            self.process_knowledge_base_file(file_path)
    
    def on_modified(self, event):
        """Handle file modification events"""
        if event.is_directory:
            return
        
        file_path = event.src_path
        file_name = os.path.basename(file_path)
        
        # Only process Excel files
        if not file_name.lower().endswith(('.xlsx', '.xls')):
            return
        
        # Check if file is A.xlsx (Knowledge Base file)
        if file_name == 'A.xlsx':
            # Avoid processing the same file multiple times in quick succession
            current_time = time.time()
            with self.processing_lock:
                if file_path in self.last_processed:
                    if current_time - self.last_processed[file_path] < 2:  # 2 seconds cooldown
                        return
                self.last_processed[file_path] = current_time
            
            # Wait a moment to ensure file is fully written
            time.sleep(1)
            
            print(f"\n{'='*60}")
            print(f"Detected modified Knowledge Base file: {file_name}")
            print(f"{'='*60}")
            self.process_knowledge_base_file(file_path)
    
    def process_knowledge_base_file(self, file_path):
        """Process Knowledge Base file and generate embeddings"""
        try:
            print(f"Processing Knowledge Base file: {file_path}")
            
            # Preprocess master file
            preprocess_master_file()
            
            # Initialize ANN index
            init_ann_index()
            
            print(f"Knowledge Base file processing completed successfully!")
            print(f"{'='*60}\n")
            
        except Exception as e:
            print(f"Error processing Knowledge Base file: {str(e)}")
            import traceback
            traceback.print_exc()
            print(f"{'='*60}\n")


def start_file_watcher():
    """Start file watcher for uploads directory"""
    try:
        # Ensure uploads directory exists
        if not os.path.exists('uploads'):
            os.makedirs('uploads')
            print("Created uploads directory for file watching")
        
        # Create observer and handler
        event_handler = FileWatcherHandler()
        observer = Observer()
        observer.schedule(event_handler, 'uploads', recursive=False)
        
        # Start observer in a separate thread
        observer.start()
        
        print("\n" + "="*60)
        print("FILE WATCHER STARTED")
        print("="*60)
        print(f"Monitoring directory: {os.path.abspath('uploads')}")
        print("Automatically processing Knowledge Base files (A.xlsx)...")
        print("="*60 + "\n")
        
        # Store observer for cleanup
        global kb_file_observer, prediction_file_observer
        file_observer = observer
        
    except Exception as e:
        print(f"Error starting file watcher: {str(e)}")
        print("File watcher not available. Manual upload through web interface still works.")